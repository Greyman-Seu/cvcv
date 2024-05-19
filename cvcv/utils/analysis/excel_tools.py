# -*- encoding: utf-8 -*-
# Copyright (c) TenStep[yangkun.zhu]. All rights reserved.

from __future__ import absolute_import
from __future__ import division
from __future__ import print_function
from __future__ import unicode_literals


try:
    import openpyxl
    from openpyxl.drawing.image import Image
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl module is not found. please install openpyxl")

__doc__ = "usage:: import cvcv.tools.analysis.excel_tools as et"

__all__ = []


def create_xlsx(title="DEMO"):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = title
    return workbook, worksheet


def create_xlsx_fromfile(path_xlsx, title=None, **kwarg):
    workbook = openpyxl.load_workbook(path_xlsx, **kwarg)
    worksheet = workbook.active
    if title is not None:
        worksheet.title = title
    return workbook, worksheet


def create_xlsx_fromlist(list_data, title="DEMO", head_row=None):
    if head_row is not None:
        assert len(head_row) == len(list_data[0])
    workbook, worksheet = create_xlsx(title)
    if head_row is not None:
        list_data.insert(0, head_row)
    for row_i in range(1, len(list_data) + 1):
        for col_i in range(1, len(list_data[0]) + 1):
            add_cell(worksheet, row_i, col_i, list_data[row_i - 1][col_i - 1])
    return workbook, worksheet


def add_cell(worksheet, row, col, val):
    cellref = worksheet.cell(row=row, column=col)
    cellref.value = val
    return worksheet


def save(workbook, path_sheet):
    """workbook, not worksheet"""
    workbook.save(path_sheet)


def set_row_h(worksheet, row_i, row_h):
    assert row_i > 0
    worksheet.row_dimensions[row_i].height = int(row_h)


def set_col_w(worksheet, col_i, col_w):
    assert col_i > 0
    worksheet.column_dimensions[get_column_letter[col_i]].width = int(col_w)


def insert_img(worksheet, img_path, row, column, imgh, imgw):
    # no test
    # ref: https://python.hotexamples.com/examples/openpyxl.drawing/Image/-/python-image-class-examples.html
    cellref = worksheet.cell(row=row, column=column)
    coordinate = cellref.coordinate
    img = Image(img_path)
    img.height = 50
    img.width = int(50 * imgw / imgh)
    img.anchor = coordinate
    worksheet.add_image(img)
    return worksheet
